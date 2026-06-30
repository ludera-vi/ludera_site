import io
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib.auth import login as auth_login, authenticate
from django.contrib.auth.models import User
from django.contrib import messages
from django.core.cache import cache
from django.db.models import Count, Max, Q, OuterRef, Subquery, Value
from django.http import JsonResponse, HttpResponse
from django.template.loader import render_to_string
from django.views.decorators.http import require_POST
from django.urls import reverse
from django.core.paginator import Paginator
from django.utils import timezone

from .models import Client, Call, ClientDocument, ClientActivity, Board, Column, Card, Chat, ChatMember, ChatMessage, InfoTopic, InfoTopicRead
from .forms import ClientForm
from users.models import ManagerSuggestion, SuggestionMessage


def manager_login(request):
    if request.user.is_authenticated:
        role = getattr(getattr(request.user, 'profile', None), 'role', 'admin')
        if role == 'manager':
            return redirect('sales:dashboard')
        return redirect('cabinet:dashboard')
    error = None
    if request.method == 'POST':
        username = request.POST.get('username', '')
        password = request.POST.get('password', '')
        user = authenticate(request, username=username, password=password)
        if user is not None:
            role = getattr(getattr(user, 'profile', None), 'role', 'admin')
            if role == 'manager' or user.is_superuser:
                auth_login(request, user)
                if role == 'manager':
                    return redirect('sales:dashboard')
                return redirect('cabinet:dashboard')
            error = 'У вас нет доступа к панели менеджера.'
        else:
            error = 'Неверный логин или пароль.'
    return render(request, 'sales/login.html', {'error': error})


def manager_logout(request):
    from django.contrib.auth import logout
    logout(request)
    return redirect('sales:login')


def _manager_required(view_func):
    @login_required(login_url='/manager/login/')
    def _wrapped(request, *args, **kwargs):
        role = getattr(getattr(request.user, 'profile', None), 'role', 'admin')
        if role != 'manager' and not request.user.is_superuser:
            return redirect('/cabinet/')
        return view_func(request, *args, **kwargs)
    _wrapped.__name__ = view_func.__name__
    return _wrapped


def _ns(request):
    return 'sales' if request.path.startswith('/manager/') else 'cabinet'


def _can_modify(request, client):
    if request.user.is_superuser:
        return True
    if client.assigned_manager is None:
        return True
    return client.assigned_manager == request.user


def _base_template(request):
    return 'sales/base.html' if request.path.startswith('/manager/') else 'cabinet/base.html'


def _list_viewname(ns, section):
    map = {
        'called': f'{ns}:called_list',
        'in_progress': f'{ns}:in_progress_list',
        'archive': f'{ns}:archive_list',
        'completed': f'{ns}:completed_list',
        'refusal': f'{ns}:refusal_list',
    }
    return reverse(map.get(section, f'{ns}:client_list'))


def _ctx(request, **kw):
    ns = _ns(request)
    kw.setdefault('base_template', _base_template(request))
    kw.setdefault('ns', ns)

    sidebar = {}
    if ns in ('sales', 'cabinet'):
        cache_key = f'sidebar_counts:{ns}:{request.user.pk}'
        sidebar = cache.get(cache_key)
        if sidebar is None:
            latest_status = Call.objects.filter(client=OuterRef('pk')).order_by('-created_at').values('status')[:1]
            qs = Client.objects.filter(is_archived=False, is_deleted=False).annotate(_last_status=Subquery(latest_status))
            if ns == 'sales' and not request.user.is_superuser:
                qs = qs.filter(assigned_manager=request.user)
            counts = dict(qs.values('_last_status').annotate(cnt=Count('id')).values_list('_last_status', 'cnt'))
            sidebar = {
                'called': counts.get('call_back', 0) + counts.get('unavailable', 0),
                'in_progress': sum(counts.get(s, 0) for s in ('negotiation', 'tz_creation', 'tz_approval', 'contract_signing', 'in_progress')),
                'completed': counts.get('completed', 0),
                'archive': Client.objects.filter(is_archived=True, is_deleted=False).count(),
                'deleted': Client.objects.filter(is_deleted=True).count(),
            }
            if ns == 'sales':
                sidebar['suggestions_unread'] = SuggestionMessage.objects.filter(
                    suggestion__manager=request.user,
                    author__is_superuser=True,
                    is_read=False
                ).count()
            from main.notification import _chat_unread_for_user
            sidebar['chat_unread'] = _chat_unread_for_user(request.user.pk)
            total_info = InfoTopic.objects.count()
            read_by_user = InfoTopicRead.objects.filter(user=request.user).values('topic').distinct().count()
            sidebar['info_unread'] = total_info - read_by_user
            cache.set(cache_key, sidebar, 10)
        kw['sidebar_counts'] = sidebar

    kw.setdefault('managers', User.objects.filter(is_active=True).order_by('last_name', 'first_name'))
    if ns == 'cabinet':
        kw.setdefault('url_clients', reverse('cabinet:client_list'))
        kw.setdefault('url_call_update', '/cabinet/calls/{pk}/update/')
        kw.setdefault('url_in_progress', reverse('cabinet:in_progress_list'))
    else:
        kw.setdefault('url_clients', reverse('sales:client_list'))
        kw.setdefault('url_call_update', '/manager/calls/{pk}/update/')
        kw.setdefault('url_in_progress', reverse('sales:in_progress_list'))
    return kw


@_manager_required
def dashboard(request):
    total_clients = Client.objects.filter(is_deleted=False).count()
    uncalled = Client.objects.filter(
        ~Q(calls__isnull=False), is_archived=False, is_deleted=False
    ).count()
    in_progress = Call.objects.exclude(status__in=['refusal', 'completed', 'call_back', 'unavailable']).count()
    archived = Client.objects.filter(is_archived=True, is_deleted=False).count()
    my_clients = Client.objects.filter(assigned_manager=request.user, is_deleted=False).count()

    status_counts = list(
        Call.objects.values('status')
        .annotate(count=Count('id'))
        .order_by('status')
    )
    total_calls = sum(s['count'] for s in status_counts) or 1

    return render(request, 'sales/dashboard.html', _ctx(request,
        title='Дашборд',
        total_clients=total_clients,
        uncalled=uncalled,
        in_progress=in_progress,
        archived=archived,
        my_clients=my_clients,
        status_counts=status_counts,
        total_calls=total_calls,
    ))


STATUS_SEARCH_MAP = {
    'согласование': 'negotiation',
    'создание тз': 'tz_creation',
    'согласование тз': 'tz_approval',
    'подписание договора': 'contract_signing',
    'в работе': 'in_progress',
    'выполнено': 'completed',
    'отказ': 'refusal',
    'перезвонить': 'call_back',
    'не доступен': 'unavailable',
    'новый': '__new__',
}

STATUS_TEXT_SEARCH = {
    'negotiation': 'Согласование',
    'tz_creation': 'Создание ТЗ',
    'tz_approval': 'Согласование ТЗ',
    'contract_signing': 'Подписание договора',
    'in_progress': 'В работе',
    'completed': 'Выполнено',
    'refusal': 'Отказ',
    'call_back': 'Перезвонить',
    'unavailable': 'Не доступен',
}


def _last_status_clients(*statuses):
    latest_status = Call.objects.filter(client=OuterRef('pk')).order_by('-created_at').values('status')[:1]
    return Client.objects.annotate(_last_status=Subquery(latest_status)).filter(_last_status__in=statuses)


def _apply_search(qs, search):
    if not search:
        return qs
    search_lower = search.lower()
    matched_status = STATUS_SEARCH_MAP.get(search_lower)
    if matched_status == '__new__':
        return qs.filter(calls__isnull=True).distinct()
    if matched_status:
        return qs.filter(calls__status=matched_status).distinct()
    status_pks = [pk for pk, txt in STATUS_TEXT_SEARCH.items() if search_lower in txt.lower()]
    q = (
        Q(phone__icontains=search) |
        Q(company_name__icontains=search) | Q(city__icontains=search) |
        Q(industry__icontains=search) | Q(comment__icontains=search) |
        Q(online_booking__icontains=search) |
        Q(website_link__icontains=search) | Q(map_link__icontains=search) |
        Q(assigned_manager__first_name__icontains=search) |
        Q(assigned_manager__last_name__icontains=search) |
        Q(assigned_manager__email__icontains=search)
    )
    if status_pks:
        q |= Q(calls__status__in=status_pks)
    return qs.filter(q).distinct()


@_manager_required
def client_list(request):
    qs = Client.objects.filter(is_archived=False, is_deleted=False).exclude(
        pk__in=_last_status_clients('refusal').values('pk')
    ).select_related('assigned_manager', 'imported_by')
    search = request.GET.get('q', '').strip()
    no_manager = request.GET.get('no_manager') == '1'
    if no_manager:
        qs = qs.filter(assigned_manager__isnull=True)
    qs = _apply_search(qs, search)

    per_page = qs.count() if request.GET.get('all') == '1' else 20
    paginator = Paginator(qs, max(per_page, 1))
    page_number = request.GET.get('page', 1)
    page_obj = paginator.get_page(page_number)
    clients = page_obj.object_list

    ctx = dict(
        clients=clients,
        page_obj=page_obj,
        search=search,
        no_manager=no_manager,
        active_section='clients',
        user=request.user,
    )

    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        rows_html = render_to_string('sales/client_list_rows.html', ctx)
        footer_html = render_to_string('sales/client_list_footer.html', ctx)
        return JsonResponse({'rows': rows_html, 'footer': footer_html})

    return render(request, 'sales/client_list.html', _ctx(request,
        title='База клиентов',
        clients=clients,
        page_obj=page_obj,
        active_section='clients',
        search=search,
        no_manager=no_manager,
    ))


@_manager_required
def called_list(request):
    qs = _last_status_clients('call_back', 'unavailable').filter(is_archived=False, is_deleted=False).select_related('assigned_manager', 'imported_by')
    if not request.user.is_superuser:
        qs = qs.filter(assigned_manager=request.user)
    search = request.GET.get('q', '').strip()
    qs = _apply_search(qs, search)
    per_page = qs.count() if request.GET.get('all') == '1' else 20
    paginator = Paginator(qs, max(per_page, 1))
    page_obj = paginator.get_page(request.GET.get('page', 1))
    ctx = dict(clients=page_obj.object_list, page_obj=page_obj, search=search, active_section='called', user=request.user)
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return JsonResponse({'rows': render_to_string('sales/client_list_rows.html', ctx), 'footer': render_to_string('sales/client_list_footer.html', ctx)})
    return render(request, 'sales/client_list.html', _ctx(request, title='Перезвонить', **ctx))


@_manager_required
def in_progress_list(request):
    qs = _last_status_clients('negotiation', 'tz_creation', 'tz_approval', 'contract_signing', 'in_progress').filter(
        is_archived=False, is_deleted=False
    ).select_related('assigned_manager', 'imported_by')
    search = request.GET.get('q', '').strip()
    qs = _apply_search(qs, search)
    per_page = qs.count() if request.GET.get('all') == '1' else 20
    paginator = Paginator(qs, max(per_page, 1))
    page_obj = paginator.get_page(request.GET.get('page', 1))
    ctx = dict(clients=page_obj.object_list, page_obj=page_obj, search=search, active_section='in_progress', user=request.user)
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return JsonResponse({'rows': render_to_string('sales/client_list_rows.html', ctx), 'footer': render_to_string('sales/client_list_footer.html', ctx)})
    return render(request, 'sales/client_list.html', _ctx(request, title='В работе', **ctx))


@_manager_required
def archive_list(request):
    qs = Client.objects.filter(is_archived=True, is_deleted=False).select_related('assigned_manager', 'imported_by')
    search = request.GET.get('q', '').strip()
    qs = _apply_search(qs, search)
    per_page = qs.count() if request.GET.get('all') == '1' else 20
    paginator = Paginator(qs, max(per_page, 1))
    page_obj = paginator.get_page(request.GET.get('page', 1))
    ctx = dict(clients=page_obj.object_list, page_obj=page_obj, search=search, active_section='archive', user=request.user)
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return JsonResponse({'rows': render_to_string('sales/client_list_rows.html', ctx), 'footer': render_to_string('sales/client_list_footer.html', ctx)})
    return render(request, 'sales/client_list.html', _ctx(request, title='Архив', **ctx))


@_manager_required
def completed_list(request):
    qs = _last_status_clients('completed').filter(
        is_archived=False, is_deleted=False
    ).select_related('assigned_manager', 'imported_by')
    search = request.GET.get('q', '').strip()
    qs = _apply_search(qs, search)
    per_page = qs.count() if request.GET.get('all') == '1' else 20
    paginator = Paginator(qs, max(per_page, 1))
    page_obj = paginator.get_page(request.GET.get('page', 1))
    ctx = dict(clients=page_obj.object_list, page_obj=page_obj, search=search, active_section='completed', user=request.user)
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return JsonResponse({'rows': render_to_string('sales/client_list_rows.html', ctx), 'footer': render_to_string('sales/client_list_footer.html', ctx)})
    return render(request, 'sales/client_list.html', _ctx(request, title='Выполненные', **ctx))


@_manager_required
def refusal_list(request):
    qs = _last_status_clients('refusal').filter(is_archived=False, is_deleted=False).select_related('assigned_manager', 'imported_by')
    search = request.GET.get('q', '').strip()
    qs = _apply_search(qs, search)
    per_page = qs.count() if request.GET.get('all') == '1' else 20
    paginator = Paginator(qs, max(per_page, 1))
    page_obj = paginator.get_page(request.GET.get('page', 1))
    ctx = dict(clients=page_obj.object_list, page_obj=page_obj, search=search, active_section='refusal', user=request.user)
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return JsonResponse({'rows': render_to_string('sales/client_list_rows.html', ctx), 'footer': render_to_string('sales/client_list_footer.html', ctx)})
    return render(request, 'sales/client_list.html', _ctx(request, title='Отказ', **ctx))


@_manager_required
@require_POST
def client_restore_from_refusal(request, pk):
    if not request.user.is_superuser:
        messages.error(request, 'Доступ запрещён')
        return redirect('sales:client_list')
    client = get_object_or_404(Client, pk=pk)
    last_call = client.calls.filter(status='refusal').last()
    if last_call:
        last_call.status = 'in_progress'
        last_call.save(update_fields=['status'])
        ClientActivity.objects.create(
            client=client, user=request.user, activity_type='status',
            title='Восстановлен из отказа',
            old_value='Отказ', new_value='В работе',
        )
        messages.success(request, f'Клиент {client.company_name or client.phone} восстановлен из отказа')
    return redirect('sales:refusal_list')


@_manager_required
def client_detail(request, pk):
    client = get_object_or_404(Client, pk=pk)
    if client.is_deleted and not request.user.is_superuser:
        messages.error(request, 'Клиент недоступен')
        return redirect('sales:client_list')
    calls = client.calls.select_related('manager').all()
    documents = client.documents.select_related('uploaded_by').all()
    activities = client.activities.select_related('user').all()
    form = ClientForm(instance=client)
    from_section = request.GET.get('from', 'clients')
    can_modify = _can_modify(request, client)
    if from_section in ('archive', 'refusal') and not request.user.is_superuser:
        can_modify = False
    return render(request, 'sales/client_detail.html', _ctx(request,
        title=f'Клиент — {client.company_name or client.phone}',
        client=client,
        calls=calls,
        documents=documents,
        activities=activities,
        form=form,
        can_modify=can_modify,
        active_section=from_section,
    ))


@_manager_required
@require_POST
def create_call(request, pk):
    client = get_object_or_404(Client, pk=pk)
    if not _can_modify(request, client):
        messages.error(request, 'Нет прав на изменение этого клиента')
        ns = _ns(request)
        return redirect('cabinet:client_list' if ns == 'cabinet' else 'sales:client_list')
    editable_fields = ['phone', 'company_name', 'city', 'industry', 'online_booking', 'comment', 'website_link', 'map_link']
    for field in editable_fields:
        if field in request.POST:
            old = str(getattr(client, field, ''))
            new = request.POST[field]
            if old != new:
                setattr(client, field, new)
                label = ClientForm._meta.model._meta.get_field(field).verbose_name or field
                model_field = ClientForm._meta.model._meta.get_field(field)
                choices = getattr(model_field, 'choices', None)
                if choices:
                    cmap = dict(choices)
                    old = cmap.get(old, old)
                    new = cmap.get(new, new)
                ClientActivity.objects.create(
                    client=client, user=request.user, activity_type='edit',
                    title=f'Изменено поле «{label}»',
                    old_value=old[:500], new_value=new[:500],
                )
    if request.user.is_superuser and 'assigned_manager' in request.POST:
        mgr_id = request.POST['assigned_manager']
        client.assigned_manager_id = int(mgr_id) if mgr_id else None
    elif not client.assigned_manager:
        client.assigned_manager = request.user
    client.save()
    status = request.POST.get('status', 'in_progress')
    if status not in dict(Call._meta.get_field('status').flatchoices):
        status = 'in_progress'
    call = Call.objects.create(client=client, manager=request.user, status=status)
    ClientActivity.objects.create(
        client=client, user=request.user, activity_type='call',
        title='Создан обзвон',
        new_value=dict(Call._meta.get_field('status').flatchoices).get(status, status),
    )
    messages.success(request, f'Обзвон создан для {client.company_name or client.phone}')
    ns = _ns(request)
    if ns == 'cabinet':
        return redirect('cabinet:client_list')
    return redirect('sales:client_list')


@_manager_required
@require_POST
def toggle_archive(request, pk):
    client = get_object_or_404(Client, pk=pk)
    if not _can_modify(request, client):
        messages.error(request, 'Нет прав на изменение этого клиента')
        return redirect('sales:client_list')
    client.is_archived = not client.is_archived
    client.save()
    status = 'архивирован' if client.is_archived else 'извлечён из архива'
    messages.success(request, f'Клиент {status}')
    ns = _ns(request)
    if ns == 'cabinet':
        return redirect('cabinet:client_list')
    return redirect('sales:client_list')


@_manager_required
@require_POST
def call_update(request, pk):
    call = get_object_or_404(Call, pk=pk)
    comment = request.POST.get('comment', '')
    status = request.POST.get('status', '')
    if status in dict(Call._meta.get_field('status').flatchoices):
        call.status = status
    call.comment = comment
    call.save()
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return JsonResponse({'ok': True})
    messages.success(request, 'Данные обновлены')
    ns = _ns(request)
    if ns == 'cabinet':
        return redirect('cabinet:client_detail', pk=call.client_id)
    return redirect('sales:client_detail', pk=call.client_id)


@_manager_required
def client_create(request):
    if request.method == 'POST':
        form = ClientForm(request.POST)
        if form.is_valid():
            client = form.save(commit=False)
            client.imported_by = request.user
            client.save()
            messages.success(request, f'Клиент {client.company_name or client.phone} создан')
            ns = _ns(request)
            if ns == 'cabinet':
                return redirect('cabinet:client_detail', pk=client.pk)
            return redirect('sales:client_detail', pk=client.pk)
    else:
        form = ClientForm()
    return render(request, 'sales/client_form.html', _ctx(request,
        title='Новый клиент',
        form=form,
    ))


@_manager_required
def client_export(request):
    if not request.user.is_superuser:
        messages.error(request, 'Доступ запрещён')
        return redirect(_ns(request) + ':client_list')
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill

    qs = Client.objects.filter(is_archived=False, is_deleted=False).select_related('assigned_manager').all()
    search = request.GET.get('q', '').strip()
    no_manager = request.GET.get('no_manager') == '1'
    if no_manager:
        qs = qs.filter(assigned_manager__isnull=True)
    qs = _apply_search(qs, search)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Клиенты'

    headers = ['Сфера', 'Город', 'Наименование', 'Телефон', 'Онлайн-запись',
               'Комментарий', 'Ответственный', 'Ссылка на сайт', 'Ссылка на Яндекс/2ГИС',
               'В архиве', 'Дата создания']
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='2D8F5E', end_color='2D8F5E', fill_type='solid')
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    for row, client in enumerate(qs, 2):
        ws.cell(row=row, column=1, value=client.industry or '')
        ws.cell(row=row, column=2, value=client.city or '')
        ws.cell(row=row, column=3, value=client.company_name or '')
        ws.cell(row=row, column=4, value=client.phone)
        ws.cell(row=row, column=5, value=client.online_booking or '')
        ws.cell(row=row, column=6, value=client.comment or '')
        ws.cell(row=row, column=7, value=client.assigned_manager.email if client.assigned_manager else '')
        ws.cell(row=row, column=8, value=client.website_link or '')
        ws.cell(row=row, column=9, value=client.map_link or '')
        ws.cell(row=row, column=10, value='Да' if client.is_archived else '')
        ws.cell(row=row, column=11, value=client.created_at.strftime('%d.%m.%Y %H:%M'))

    for col in range(1, len(headers) + 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 20

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="clients.xlsx"'
    wb.save(response)
    return response


@_manager_required
def client_import(request):
    if not request.user.is_superuser:
        messages.error(request, 'Доступ запрещён')
        return redirect(_ns(request) + ':client_list')
    if request.method == 'POST' and request.FILES.get('file'):
        import openpyxl
        import csv
        import os

        file = request.FILES['file']
        ext = os.path.splitext(file.name)[1].lower()
        rows = []

        clients_url = 'cabinet:client_list' if _ns(request) == 'cabinet' else 'sales:client_list'

        if ext in ('.xlsx', '.xls'):
            try:
                wb = openpyxl.load_workbook(file)
                ws = wb.active
                rows = list(ws.iter_rows(min_row=2, values_only=True))
            except Exception:
                messages.error(request, 'Не удалось прочитать файл. Загрузите .xlsx')
                return redirect(clients_url)
        elif ext == '.csv':
            try:
                decoded = file.read().decode('utf-8-sig')
                reader = csv.reader(io.StringIO(decoded))
                next(reader, None)
                rows = list(reader)
            except Exception:
                messages.error(request, 'Не удалось прочитать CSV. Проверьте кодировку (UTF-8)')
                return redirect(clients_url)
        else:
            messages.error(request, 'Поддерживаются только .xlsx, .xls и .csv')
            return redirect(clients_url)

        if not rows:
            messages.error(request, 'Файл пуст')
            return redirect(clients_url)

        created = 0
        updated = 0
        skipped = 0
        for row in rows:
            industry = str(row[0]).strip() if len(row) >= 1 and row[0] else ''
            city = str(row[1]).strip() if len(row) >= 2 and row[1] else ''
            company = str(row[2]).strip() if len(row) >= 3 and row[2] else ''
            phone = str(row[3]).strip() if len(row) >= 4 and row[3] else None
            raw_online = str(row[4]).strip() if len(row) >= 5 and row[4] else ''
            online_booking = raw_online if raw_online.startswith('http') else ''
            comment = str(row[5]).strip() if len(row) >= 6 and row[5] else ''
            raw_website = str(row[7]).strip() if len(row) >= 8 and row[7] else ''
            website_link = raw_website if raw_website.startswith('http') else ''
            raw_map = str(row[8]).strip() if len(row) >= 9 and row[8] else ''
            map_link = raw_map if raw_map.startswith('http') else ''

            if not phone:
                skipped += 1
                continue

            existing = Client.objects.filter(phone=phone).first()
            if existing:
                changed = False
                for field, val in [('industry', industry), ('city', city),
                                   ('company_name', company),
                                   ('online_booking', online_booking),
                                   ('comment', comment),
                                   ('website_link', website_link),
                                   ('map_link', map_link)]:
                    if val and getattr(existing, field) != val:
                        setattr(existing, field, val)
                        changed = True
                if changed:
                    existing.save()
                    updated += 1
                else:
                    skipped += 1
            else:
                Client.objects.create(
                    industry=industry, city=city, company_name=company,
                    phone=phone,
                    online_booking=online_booking, comment=comment,
                    website_link=website_link, map_link=map_link,
                    imported_by=request.user,
                )
                created += 1

        parts = []
        if created:
            parts.append(f'добавлено {created}')
        if updated:
            parts.append(f'обновлено {updated}')
        if skipped:
            parts.append(f'пропущено {skipped}')
        messages.success(request, 'Импорт завершён. ' + ', '.join(parts))
        return redirect(clients_url)

    return render(request, 'sales/client_import.html', _ctx(request,
        title='Импорт клиентов',
    ))


@_manager_required
def client_update(request, pk):
    client = get_object_or_404(Client, pk=pk)
    from_section = request.POST.get('from_section', request.GET.get('from', 'clients'))
    if request.method == 'POST':
        if not _can_modify(request, client):
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({'ok': False, 'error': 'no permission'}, status=403)
            messages.error(request, 'Нет прав на изменение этого клиента')
            return redirect('sales:client_list')

        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            editable_fields = ['phone', 'company_name', 'city', 'industry', 'online_booking', 'comment', 'website_link', 'map_link']
            for field in editable_fields:
                if field in request.POST:
                    old = str(getattr(client, field, ''))
                    new = request.POST[field]
                    if old != new:
                        setattr(client, field, new)
                        label = ClientForm._meta.model._meta.get_field(field).verbose_name or field
                        model_field = ClientForm._meta.model._meta.get_field(field)
                        choices = getattr(model_field, 'choices', None)
                        if choices:
                            cmap = dict(choices)
                            old = cmap.get(old, old)
                            new = cmap.get(new, new)
                        ClientActivity.objects.create(
                            client=client, user=request.user, activity_type='edit',
                            title=f'Изменено поле «{label}»',
                            old_value=old[:500], new_value=new[:500],
                        )
            if request.user.is_superuser and 'assigned_manager' in request.POST:
                mgr_id = request.POST['assigned_manager']
                new_mgr_id = int(mgr_id) if mgr_id else None
                if client.assigned_manager_id != new_mgr_id:
                    old_name = str(client.assigned_manager) if client.assigned_manager else '—'
                    client.assigned_manager_id = new_mgr_id
                    new_name = str(User.objects.get(pk=new_mgr_id)) if new_mgr_id else '—'
                    ClientActivity.objects.create(
                        client=client, user=request.user, activity_type='edit',
                        title='Изменён ответственный',
                        old_value=old_name, new_value=new_name,
                    )
            client.save()
            return JsonResponse({'ok': True})

        form = ClientForm(request.POST, instance=client)
        if not form.is_valid():
            for field, errors in form.errors.items():
                for err in errors:
                    messages.error(request, f'{form.fields[field].label or field}: {err}')
            return redirect(_list_viewname(_ns(request), from_section))

        changes = []
        for field in form.changed_data:
            old = str(form.initial.get(field, ''))
            new = str(form.cleaned_data.get(field, ''))
            if old != new:
                label = form.fields[field].label if field in form.fields else field
                changes.append((field, label, old, new))
        form.save()
        if request.user.is_superuser and 'assigned_manager' in request.POST:
            mgr_id = request.POST['assigned_manager']
            new_mgr_id = int(mgr_id) if mgr_id else None
            if client.assigned_manager_id != new_mgr_id:
                old_name = str(client.assigned_manager) if client.assigned_manager else '—'
                client.assigned_manager_id = new_mgr_id
                client.save(update_fields=['assigned_manager'])
                new_name = str(User.objects.get(pk=new_mgr_id)) if new_mgr_id else '—'
                ClientActivity.objects.create(
                    client=client, user=request.user, activity_type='edit',
                    title='Изменён ответственный',
                    old_value=old_name, new_value=new_name,
                )
        for field, label, old, new in changes:
            choices = getattr(form.fields.get(field), 'choices', None)
            if choices:
                cmap = dict(choices)
                old = cmap.get(old, old)
                new = cmap.get(new, new)
            ClientActivity.objects.create(
                client=client, user=request.user, activity_type='edit',
                title=f'Изменено поле «{label}»',
                old_value=old[:500], new_value=new[:500],
            )

        doc_file = request.FILES.get('doc_file')
        if doc_file:
            title = doc_file.name
            ClientDocument.objects.create(
                client=client, title=title, file=doc_file, uploaded_by=request.user,
            )
            ClientActivity.objects.create(
                client=client, user=request.user, activity_type='document',
                title=f'Загружен документ «{title}»',
            )
            messages.success(request, f'Документ "{title}" загружен')

        status = request.POST.get('status', '')
        last_call = client.calls.last()
        if last_call and status in dict(Call._meta.get_field('status').flatchoices):
            old_status = last_call.status
            if old_status != status:
                last_call.status = status
                last_call.save(update_fields=['status'])
                old_label = dict(Call._meta.get_field('status').flatchoices).get(old_status, old_status)
                new_label = dict(Call._meta.get_field('status').flatchoices).get(status, status)
                ClientActivity.objects.create(
                    client=client, user=request.user, activity_type='status',
                    title='Смена статуса',
                    old_value=old_label, new_value=new_label,
                )

        messages.success(request, 'Сохранено')
        ns = _ns(request)
        return redirect(_list_viewname(ns, from_section))
    else:
        form = ClientForm(instance=client)
    return render(request, 'sales/client_form.html', _ctx(request,
        title='Редактирование клиента',
        form=form, edit_mode=True,
    ))


@_manager_required
@require_POST
def client_delete(request, pk):
    client = get_object_or_404(Client, pk=pk)
    if not _can_modify(request, client):
        messages.error(request, 'Нет прав на удаление этого клиента')
        return redirect('sales:client_list')
    if client.is_deleted:
        if not request.user.is_superuser:
            messages.error(request, 'Только администратор может удалять окончательно')
            return redirect('sales:deleted_list')
        name = str(client)
        client.delete()
        messages.success(request, f'Клиент {name} окончательно удалён')
        return redirect('sales:deleted_list')
    client.is_deleted = True
    client.save(update_fields=['is_deleted'])
    messages.success(request, f'Клиент перемещён в удалённые')
    ns = _ns(request)
    return redirect('cabinet:client_list' if ns == 'cabinet' else 'sales:client_list')


@_manager_required
def deleted_list(request):
    if not request.user.is_superuser:
        messages.error(request, 'Доступ запрещён')
        return redirect('sales:client_list')
    qs = Client.objects.filter(is_deleted=True).select_related('assigned_manager', 'imported_by')
    search = request.GET.get('q', '').strip()
    qs = _apply_search(qs, search)
    per_page = qs.count() if request.GET.get('all') == '1' else 20
    paginator = Paginator(qs, max(per_page, 1))
    page_obj = paginator.get_page(request.GET.get('page', 1))
    ctx = dict(clients=page_obj.object_list, page_obj=page_obj, search=search, active_section='deleted', user=request.user)
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return JsonResponse({'rows': render_to_string('sales/client_list_rows.html', ctx), 'footer': render_to_string('sales/client_list_footer.html', ctx)})
    return render(request, 'sales/client_list.html', _ctx(request, title='Удалённые', **ctx))


@_manager_required
@require_POST
def client_restore(request, pk):
    if not request.user.is_superuser:
        messages.error(request, 'Доступ запрещён')
        return redirect('sales:client_list')
    client = get_object_or_404(Client, pk=pk, is_deleted=True)
    client.is_deleted = False
    client.save(update_fields=['is_deleted'])
    messages.success(request, f'Клиент восстановлен')
    return redirect('sales:deleted_list')


@_manager_required
@require_POST
def delete_document(request, client_pk, doc_pk):
    document = get_object_or_404(ClientDocument, pk=doc_pk, client_id=client_pk)
    client = document.client
    if not _can_modify(request, client):
        messages.error(request, 'Нет прав на удаление документа')
        return redirect('sales:client_list')
    title = document.title
    ClientActivity.objects.create(
        client=client, user=request.user, activity_type='document',
        title=f'Удалён документ «{title}»',
    )
    document.delete()
    messages.success(request, f'Документ "{title}" удалён')
    ns = _ns(request)
    from_section = request.POST.get('from_section', 'clients')
    detail_url = 'sales:client_detail' if ns == 'sales' else 'cabinet:client_detail'
    return redirect(f'{reverse(detail_url, args=[client_pk])}?from={from_section}')


@_manager_required
def suggestion_list(request):
    open_pk = ''
    if request.method == 'POST':
        action = request.POST.get('action', '')
        is_ajax = request.headers.get('X-Requested-With') == 'XMLHttpRequest'

        if action == 'create':
            message = request.POST.get('message', '').strip()
            if message:
                sug = ManagerSuggestion.objects.create(manager=request.user, message=message)
                msg = SuggestionMessage.objects.create(suggestion=sug, author=request.user, message=message)
                from main.notification import broadcast_suggestion_message
                broadcast_suggestion_message(msg)
                messages.success(request, 'Предложение отправлено')
            else:
                messages.error(request, 'Напишите текст предложения')
        elif action == 'reply':
            pk = request.POST.get('pk', '')
            text = request.POST.get('message', '').strip()
            if pk and text:
                sug = get_object_or_404(ManagerSuggestion, pk=pk, manager=request.user)
                if not sug.is_closed:
                    msg = SuggestionMessage.objects.create(
                        suggestion=sug, author=request.user, message=text
                    )
                    from main.notification import broadcast_suggestion_message
                    broadcast_suggestion_message(msg)
                    if is_ajax:
                        from django.utils import timezone
                        t = timezone.localtime(msg.created_at)
                        return JsonResponse({
                            'ok': True,
                            'message': msg.message,
                            'author': msg.author.get_full_name() or msg.author.email,
                            'is_superuser': msg.author.is_superuser,
                            'time': t.strftime('%H:%M'),
                            'pk': msg.pk,
                        })
            open_pk = pk

        if is_ajax:
            return JsonResponse({'ok': False}, status=400)
        url = reverse('sales:suggestion_list')
        if open_pk:
            url += f'?open={open_pk}'
        return redirect(url)

    suggestions = ManagerSuggestion.objects.filter(
        manager=request.user
    ).select_related('admin').prefetch_related('messages', 'messages__author').all()

    # Compute unread for each suggestion
    suggestions_list = list(suggestions)
    for s in suggestions_list:
        s.has_unread = any(
            not m.is_read and m.author.is_superuser
            for m in s.messages.all()
        ) or s.status == 'unread'

    return render(request, 'sales/suggestion_list.html', _ctx(request,
        title='Обратная связь',
        suggestions=suggestions_list,
        active_section='suggestions',
    ))


# ─── Kanban ─────────────────────────────────────────────────────────


@_manager_required
def kanban_list(request):
    if not request.user.is_superuser:
        return render(request, 'sales/kanban_list.html', _ctx(request,
            title='Канбан',
            active_section='kanban',
            kanban_placeholder=True,
        ))
    boards = Board.objects.filter().select_related('created_by')
    return render(request, 'sales/kanban_list.html', _ctx(request,
        title='Канбан доски',
        boards=boards,
        active_section='kanban',
    ))


@_manager_required
def kanban_board(request, pk):
    board = get_object_or_404(Board, pk=pk)
    columns = list(board.columns.prefetch_related(
        'cards', 'cards__client', 'cards__responsible', 'cards__created_by'
    ).all())

    # Auto-create cards for clients in "in progress" statuses that don't have cards on this board
    existing_client_ids = set()
    for col in columns:
        for c in col.cards.all():
            if c.client_id:
                existing_client_ids.add(c.client_id)
    in_progress_clients = _last_status_clients('negotiation', 'tz_creation', 'tz_approval', 'contract_signing', 'in_progress').filter(
        is_archived=False, is_deleted=False
    ).exclude(pk__in=existing_client_ids)
    if in_progress_clients.exists() and columns:
        from main.notification import broadcast_kanban_card
        target_col = columns[0]
        for client in in_progress_clients:
            card = Card.objects.create(
                column=target_col, title=str(client),
                client=client, created_by=request.user,
            )
            broadcast_kanban_card(card, 'created')
        # Refresh columns to include new cards
        columns = list(board.columns.prefetch_related(
            'cards', 'cards__client', 'cards__responsible', 'cards__created_by'
        ).all())

    client_filter = request.GET.get('client', '').strip()
    if client_filter and client_filter.isdigit():
        for col in columns:
            col.filtered_cards = [c for c in col.cards.all() if c.client_id == int(client_filter)]
    else:
        for col in columns:
            col.filtered_cards = list(col.cards.all())
    return render(request, 'sales/kanban_board.html', _ctx(request,
        title=board.title,
        board=board,
        columns=columns,
        client_filter=client_filter,
        clients=Client.objects.filter(is_deleted=False).order_by('name', 'company_name'),
        managers=User.objects.filter(is_active=True).exclude(pk=request.user.pk).order_by('email'),
        active_section='kanban',
    ))


@_manager_required
@require_POST
def kanban_create_board(request):
    if not request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return JsonResponse({'ok': False}, status=400)
    title = request.POST.get('title', '').strip()
    if not title:
        return JsonResponse({'ok': False, 'error': 'Название обязательно'}, status=400)
    board = Board.objects.create(title=title, description=request.POST.get('description', ''), created_by=request.user)
    Column.objects.create(board=board, title='Нужно сделать', position=0)
    Column.objects.create(board=board, title='В работе', position=1)
    Column.objects.create(board=board, title='Готово', position=2)
    return JsonResponse({'ok': True, 'pk': board.pk, 'title': board.title})


@_manager_required
@require_POST
def kanban_delete_board(request, pk):
    board = get_object_or_404(Board, pk=pk)
    if not request.user.is_superuser:
        return JsonResponse({'ok': False, 'error': 'Только админ'}, status=403)
    board.delete()
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return JsonResponse({'ok': True})
    messages.success(request, 'Доска удалена')
    return redirect('sales:kanban_list')


@_manager_required
@require_POST
def kanban_create_column(request):
    if not request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return JsonResponse({'ok': False}, status=400)
    board_id = request.POST.get('board_id', '')
    title = request.POST.get('title', '').strip()
    if not board_id or not title:
        return JsonResponse({'ok': False, 'error': 'Обязательные поля'}, status=400)
    board = get_object_or_404(Board, pk=board_id)
    if board.columns.count() >= 10:
        return JsonResponse({'ok': False, 'error': 'Максимум 10 колонок'}, status=400)
    max_pos = board.columns.aggregate(m=models.Max('position'))['m'] or -1
    col = Column.objects.create(board=board, title=title, position=max_pos + 1)
    return JsonResponse({'ok': True, 'pk': col.pk, 'title': col.title, 'position': col.position})


@_manager_required
@require_POST
def kanban_edit_column(request, pk):
    if not request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return JsonResponse({'ok': False}, status=400)
    col = get_object_or_404(Column, pk=pk)
    title = request.POST.get('title', '').strip()
    if not title:
        return JsonResponse({'ok': False, 'error': 'Название обязательно'}, status=400)
    col.title = title
    col.save(update_fields=['title'])
    return JsonResponse({'ok': True})


@_manager_required
@require_POST
def kanban_delete_column(request, pk):
    col = get_object_or_404(Column, pk=pk)
    board_pk = col.board_id
    col.delete()
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return JsonResponse({'ok': True})
    ns = _ns(request)
    return redirect(f'{ns}:kanban_board', pk=board_pk)


@_manager_required
@require_POST
def kanban_create_card(request):
    if not request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return JsonResponse({'ok': False}, status=400)
    column_id = request.POST.get('column_id', '')
    title = request.POST.get('title', '').strip()
    if not column_id or not title:
        return JsonResponse({'ok': False, 'error': 'Обязательные поля'}, status=400)
    col = get_object_or_404(Column, pk=column_id)
    client_id = request.POST.get('client_id', '').strip()
    responsible_id = request.POST.get('responsible_id', '').strip()
    client = None
    responsible = None
    if client_id:
        client = get_object_or_404(Client, pk=client_id)
    if responsible_id:
        responsible = get_object_or_404(User, pk=responsible_id)
    max_pos = col.cards.aggregate(m=models.Max('position'))['m'] or -1
    card = Card.objects.create(
        column=col, title=title,
        description=request.POST.get('description', ''),
        client=client, responsible=responsible,
        created_by=request.user, position=max_pos + 1,
    )
    from main.notification import broadcast_kanban_card
    broadcast_kanban_card(card, 'created')
    return JsonResponse({
        'ok': True, 'pk': card.pk, 'title': card.title,
        'client_name': str(card.client) if card.client else '',
        'responsible_name': card.responsible.get_full_name() or card.responsible.email if card.responsible else '',
    })


@_manager_required
def kanban_card_detail(request, pk):
    card = get_object_or_404(Card.objects.select_related('column__board', 'client', 'responsible', 'created_by'), pk=pk)
    can_edit = request.user.is_superuser or card.responsible == request.user
    if request.method == 'POST' and can_edit:
        title = request.POST.get('title', '').strip()
        if not title:
            messages.error(request, 'Название обязательно')
        else:
            card.title = title
            card.description = request.POST.get('description', '')
            client_id = request.POST.get('client_id', '').strip()
            responsible_id = request.POST.get('responsible_id', '').strip()
            card.client = get_object_or_404(Client, pk=client_id) if client_id else None
            if responsible_id and request.user.is_superuser:
                card.responsible = get_object_or_404(User, pk=responsible_id)
            elif not responsible_id:
                card.responsible = None
            card.save()
            from main.notification import broadcast_kanban_card
            broadcast_kanban_card(card, 'updated')
            messages.success(request, 'Карточка обновлена')
            ns = _ns(request)
            return redirect(f'{ns}:kanban_board', pk=card.column.board_id)
    return render(request, 'sales/kanban_card_detail.html', _ctx(request,
        title=card.title,
        card=card,
        can_edit=can_edit,
        clients=Client.objects.filter(is_deleted=False).order_by('name', 'company_name'),
        managers=User.objects.filter(is_active=True).exclude(pk=request.user.pk).order_by('email'),
        active_section='kanban',
    ))


@_manager_required
@require_POST
def kanban_delete_card(request, pk):
    card = get_object_or_404(Card, pk=pk)
    if not request.user.is_superuser and card.responsible != request.user:
        return JsonResponse({'ok': False, 'error': 'Нет прав'}, status=403)
    board_pk = card.column.board_id
    from main.notification import broadcast_kanban_card
    broadcast_kanban_card(card, 'deleted')
    card.delete()
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return JsonResponse({'ok': True})
    messages.success(request, 'Карточка удалена')
    ns = _ns(request)
    return redirect(f'{ns}:kanban_board', pk=board_pk)


@_manager_required
@require_POST
def kanban_move_card(request, pk):
    if not request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return JsonResponse({'ok': False}, status=400)
    card = get_object_or_404(Card, pk=pk)
    if not request.user.is_superuser and card.responsible != request.user:
        return JsonResponse({'ok': False, 'error': 'Нет прав'}, status=403)
    column_id = request.POST.get('column_id', '').strip()
    position = request.POST.get('position', '').strip()
    if not column_id:
        return JsonResponse({'ok': False, 'error': 'column_id required'}, status=400)
    new_col = get_object_or_404(Column, pk=column_id)
    old_col_id = card.column_id
    card.column = new_col
    if position:
        card.position = int(position)
    card.save(update_fields=['column', 'position'])

    # Bidirectional sync: if card has a client, update call status
    if card.client:
        col_title = new_col.title.lower()
        status_map = {
            'нужно сделать': None,
            'в работе': 'in_progress',
            'согласование': 'negotiation',
            'создание тз': 'tz_creation',
            'согласование тз': 'tz_approval',
            'подписание договора': 'contract_signing',
            'выполнено': 'completed',
            'отказ': 'refusal',
            'перезвонить': 'call_back',
            'не доступен': 'unavailable',
        }
        for key, status in status_map.items():
            if key in col_title:
                if status:
                    last_call = card.client.calls.order_by('-created_at').first()
                    if last_call:
                        last_call.status = status
                        last_call.save(update_fields=['status'])
                break

    from main.notification import broadcast_kanban_card
    broadcast_kanban_card(card, 'moved', old_col_id=old_col_id)
    return JsonResponse({'ok': True, 'new_column_id': int(column_id), 'position': card.position})


@_manager_required
@require_POST
def kanban_reorder_columns(request):
    if not request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return JsonResponse({'ok': False}, status=400)
    order = request.POST.getlist('order[]')
    for i, col_id in enumerate(order):
        Column.objects.filter(pk=col_id).update(position=i)
    return JsonResponse({'ok': True})


@_manager_required
@require_POST
def kanban_toggle_column(request, pk):
    if not request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return JsonResponse({'ok': False}, status=400)
    col = get_object_or_404(Column, pk=pk)
    col.is_collapsed = not col.is_collapsed
    col.save(update_fields=['is_collapsed'])
    return JsonResponse({'ok': True, 'is_collapsed': col.is_collapsed})


@_manager_required
def kanban_clients_json(request):
    q = request.GET.get('q', '').strip()
    qs = Client.objects.filter(is_deleted=False)
    if q:
        qs = qs.filter(Q(name__icontains=q) | Q(company_name__icontains=q) | Q(phone__icontains=q))
    results = [{'id': c.pk, 'text': str(c)} for c in qs[:20]]
    return JsonResponse({'results': results})


# ─── Chats ─────────────────────────────────────────────────────────


@_manager_required
def chat_list(request):
    qs = Chat.objects.filter(
        members__user=request.user, members__is_active=True, is_deleted=False
    ).select_related('client', 'card', 'created_by').prefetch_related('members__user').distinct()
    for chat in qs:
        last_msg = chat.messages.order_by('-created_at').first()
        chat.last_message = last_msg
        last_read = ChatMember.objects.filter(chat=chat, user=request.user).values_list('last_read_at', flat=True).first()
        q = ChatMessage.objects.filter(chat=chat).exclude(author=request.user)
        if last_read:
            q = q.filter(created_at__gt=last_read)
        chat.unread = q.count()
        if not chat.title:
            other = chat.members.exclude(user=request.user).select_related('user').first()
            if other:
                chat.display_title = other.user.get_full_name() or other.user.email
            else:
                chat.display_title = 'Чат'
        else:
            chat.display_title = chat.title
    users_qs = User.objects.filter(is_active=True).exclude(pk=request.user.pk).order_by('email')
    return render(request, 'sales/chat_list.html', _ctx(request,
        title='Чаты',
        chats=qs,
        users=users_qs,
        active_section='chats',
    ))


@_manager_required
@require_POST
def chat_create(request):
    if not request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return JsonResponse({'ok': False}, status=400)
    chat_type = request.POST.get('chat_type', 'personal')
    title = request.POST.get('title', '').strip()
    member_ids = request.POST.getlist('member_ids')
    client_id = request.POST.get('client_id', '').strip()
    card_id = request.POST.get('card_id', '').strip()
    if chat_type == 'personal' and not member_ids:
        return JsonResponse({'ok': False, 'error': 'Выберите участника'}, status=400)
    client = get_object_or_404(Client, pk=client_id) if client_id else None
    card = get_object_or_404(Card, pk=card_id) if card_id else None
    chat = Chat.objects.create(
        title=title, chat_type=chat_type,
        client=client, card=card, created_by=request.user,
    )
    ChatMember.objects.create(chat=chat, user=request.user, added_by=request.user)
    for uid in member_ids:
        u = get_object_or_404(User, pk=uid)
        ChatMember.objects.create(chat=chat, user=u, added_by=request.user)
    if chat_type == 'client' and client and client.assigned_manager:
        ChatMember.objects.get_or_create(chat=chat, user=client.assigned_manager, defaults={'added_by': request.user})
        for admin in User.objects.filter(is_superuser=True).exclude(pk=request.user.pk):
            ChatMember.objects.get_or_create(chat=chat, user=admin, defaults={'added_by': request.user})
    return JsonResponse({'ok': True, 'chat_id': chat.pk})


@_manager_required
@require_POST
def chat_send(request, pk):
    chat = get_object_or_404(Chat, pk=pk)
    if not chat.members.filter(user=request.user, is_active=True).exists():
        return JsonResponse({'ok': False, 'error': 'Нет доступа'}, status=403)
    message = request.POST.get('message', '').strip()
    if not message:
        return JsonResponse({'ok': False, 'error': 'Сообщение пустое'}, status=400)
    msg = ChatMessage.objects.create(chat=chat, author=request.user, message=message)
    ChatMember.objects.filter(chat=chat, user=request.user).update(last_read_at=msg.created_at)
    from main.notification import broadcast_chat_message
    broadcast_chat_message(msg)
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return JsonResponse({
            'ok': True, 'pk': msg.pk, 'message': msg.message,
            'author_name': msg.author.get_full_name() or msg.author.email,
            'author_id': msg.author_id,
            'is_mine': True,
            'created_at': timezone.localtime(msg.created_at).strftime('%H:%M'),
        })
    ns = _ns(request)
    return redirect(f'{ns}:chat_detail', pk=chat.pk)


@_manager_required
def chat_detail(request, pk):
    chat = get_object_or_404(Chat.objects.select_related('client', 'card'), pk=pk)
    if not chat.members.filter(user=request.user, is_active=True).exists() and not request.user.is_superuser:
        messages.error(request, 'Нет доступа')
        return redirect('cabinet:chat_list')
    # AJAX POST: quick mark-read without rendering the page
    if request.method == 'POST' and request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        ChatMember.objects.filter(chat=chat, user=request.user).update(last_read_at=timezone.now())
        from main.notification import broadcast_chat_unread_update
        broadcast_chat_unread_update(request.user)
        return JsonResponse({'ok': True})
    messages_qs = chat.messages.select_related('author').order_by('created_at')
    ChatMember.objects.filter(chat=chat, user=request.user).update(last_read_at=timezone.now())
    from main.notification import broadcast_chat_unread_update
    broadcast_chat_unread_update(request.user, chat_id=chat.pk)
    members = chat.members.filter(is_active=True).select_related('user')
    users_qs = User.objects.filter(is_active=True).exclude(pk=request.user.pk).order_by('email')
    return render(request, 'sales/chat_detail.html', _ctx(request,
        title=chat.title or 'Чат',
        chat=chat,
        chat_messages=messages_qs,
        members=members,
        users=users_qs,
        active_section='chats',
    ))


@_manager_required
def chat_messages_json(request, pk):
    chat = get_object_or_404(Chat, pk=pk)
    if not chat.members.filter(user=request.user, is_active=True).exists():
        return JsonResponse({'ok': False}, status=403)
    after = request.GET.get('after', '')
    qs = chat.messages.select_related('author').order_by('created_at')
    if after:
        qs = qs.filter(pk__gt=after)
    data = [{
        'pk': m.pk, 'message': m.message,
        'author_name': m.author.get_full_name() or m.author.email if m.author else '—',
        'author_id': m.author_id,
        'is_mine': m.author_id == request.user.pk,
        'created_at': timezone.localtime(m.created_at).strftime('%H:%M'),
    } for m in qs]
    ChatMember.objects.filter(chat=chat, user=request.user).update(last_read_at=timezone.now())
    return JsonResponse({'ok': True, 'messages': data})


@_manager_required
def chat_unread_count(request):
    qs = Chat.objects.filter(members__user=request.user, members__is_active=True, is_deleted=False)
    total = 0
    for chat in qs:
        total += ChatMessage.objects.filter(
            chat=chat, created_at__gt=chat.members.filter(user=request.user).values('last_read_at')[:1]
        ).exclude(author=request.user).count()
    return JsonResponse({'count': total})


@_manager_required
@require_POST
def chat_add_member(request, pk):
    chat = get_object_or_404(Chat, pk=pk)
    if not request.user.is_superuser and chat.created_by != request.user:
        return JsonResponse({'ok': False, 'error': 'Нет прав'}, status=403)
    user_id = request.POST.get('user_id', '').strip()
    if not user_id:
        return JsonResponse({'ok': False, 'error': 'user_id required'}, status=400)
    user = get_object_or_404(User, pk=user_id)
    member, created = ChatMember.objects.get_or_create(
        chat=chat, user=user,
        defaults={'added_by': request.user}
    )
    if not created and not member.is_active:
        member.is_active = True
        member.save(update_fields=['is_active'])
    from main.notification import broadcast_chat_member_update
    broadcast_chat_member_update(chat, 'added', user, added_by=request.user)
    return JsonResponse({'ok': True})


@_manager_required
@require_POST
def chat_remove_member(request, pk):
    chat = get_object_or_404(Chat, pk=pk)
    user_id = request.POST.get('user_id', '').strip()
    if not user_id:
        messages.error(request, 'Не указан пользователь')
        return redirect(f'{_ns(request)}:chat_detail', pk=pk)
    if not request.user.is_superuser and request.user.pk != int(user_id):
        messages.error(request, 'Нет прав')
        return redirect(f'{_ns(request)}:chat_detail', pk=pk)
    try:
        member = ChatMember.objects.get(chat=chat, user_id=user_id, is_active=True)
    except ChatMember.DoesNotExist:
        messages.error(request, 'Участник не найден')
        return redirect(f'{_ns(request)}:chat_detail', pk=pk)
    if request.user.is_superuser or member.user == request.user:
        if request.user.is_superuser:
            member.is_active = False
            member.save(update_fields=['is_active'])
        else:
            member.delete()
        from main.notification import broadcast_chat_member_update
        broadcast_chat_member_update(chat, 'removed', member.user)
        if member.user == request.user:
            return redirect(f'{_ns(request)}:chat_list')
        return redirect(f'{_ns(request)}:chat_detail', pk=pk)
    messages.error(request, 'Нет прав')
    return redirect(f'{_ns(request)}:chat_detail', pk=pk)


@_manager_required
@require_POST
def chat_delete(request, pk):
    chat = get_object_or_404(Chat, pk=pk)
    if request.user.is_superuser:
        chat.is_deleted = True
        chat.save(update_fields=['is_deleted'])
        messages.success(request, 'Чат удалён')
    else:
        member = get_object_or_404(ChatMember, chat=chat, user=request.user)
        member.delete()
        messages.success(request, 'Вы вышли из чата')
    return redirect('cabinet:chat_list')


# ─── Info Topics ──────────────────────────────────────────────────


@_manager_required
def info_list(request):
    qs = InfoTopic.objects.all()
    search = request.GET.get('q', '').strip()
    if search:
        qs = qs.filter(Q(title__icontains=search) | Q(content__icontains=search))
    read_pks = set(InfoTopicRead.objects.filter(user=request.user).values_list('topic_id', flat=True))
    for t in qs:
        t.is_read = t.pk in read_pks
    return render(request, 'sales/info_list.html', _ctx(request,
        title='Информация',
        topics=qs,
        search=search,
        active_section='info',
    ))


@_manager_required
def info_detail(request, pk):
    topic = get_object_or_404(InfoTopic, pk=pk)
    InfoTopicRead.objects.get_or_create(topic=topic, user=request.user)
    from main.notification import broadcast_info_unread_update
    broadcast_info_unread_update(request.user)
    return render(request, 'sales/info_detail.html', _ctx(request,
        title=topic.title,
        topic=topic,
        active_section='info',
    ))


@_manager_required
def info_create(request):
    if not request.user.is_superuser:
        messages.error(request, 'Только администратор может создавать топики')
        return redirect(f'{_ns(request)}:info_list')
    if request.method == 'POST':
        title = request.POST.get('title', '').strip()
        content = request.POST.get('content', '').strip()
        if not title:
            messages.error(request, 'Заголовок обязателен')
        else:
            topic = InfoTopic.objects.create(title=title, content=content, created_by=request.user)
            make_read = InfoTopicRead.objects.create(topic=topic, user=request.user)
            from main.notification import broadcast_info_topic
            broadcast_info_topic(topic, 'created')
            messages.success(request, 'Топик создан')
            return redirect(f'{_ns(request)}:info_detail', pk=topic.pk)
    return render(request, 'sales/info_list.html', _ctx(request,
        title='Новый топик',
        creating=True,
        active_section='info',
    ))


@_manager_required
def info_edit(request, pk):
    if not request.user.is_superuser:
        messages.error(request, 'Только администратор может редактировать топики')
        return redirect(f'{_ns(request)}:info_detail', pk=pk)
    topic = get_object_or_404(InfoTopic, pk=pk)
    if request.method == 'POST':
        title = request.POST.get('title', '').strip()
        content = request.POST.get('content', '').strip()
        if not title:
            messages.error(request, 'Заголовок обязателен')
        else:
            topic.title = title
            topic.content = content
            topic.save(update_fields=['title', 'content'])
            from main.notification import broadcast_info_topic
            broadcast_info_topic(topic, 'updated')
            messages.success(request, 'Топик обновлён')
            return redirect(f'{_ns(request)}:info_detail', pk=topic.pk)
    return render(request, 'sales/info_form.html', _ctx(request,
        title='Редактировать топик',
        topic=topic,
        active_section='info',
    ))


@_manager_required
@require_POST
def info_delete(request, pk):
    if not request.user.is_superuser:
        messages.error(request, 'Только администратор может удалять топики')
        return redirect(f'{_ns(request)}:info_list')
    topic = get_object_or_404(InfoTopic, pk=pk)
    topic.delete()
    from main.notification import broadcast_info_unread_update
    broadcast_info_unread_update(request.user)
    messages.success(request, 'Топик удалён')
    return redirect(f'{_ns(request)}:info_list')
